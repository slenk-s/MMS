"""
Excel 导出模块
支持将物料数据、出入库记录、盘点记录等导出为 Excel 文件
"""
from datetime import datetime
from typing import List, Dict, Optional
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelExporter:
    """Excel 导出器"""

    # 预定义样式
    HEADER_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="FFFFFF")
    HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    DATA_FONT = Font(name="Microsoft YaHei", size=11)
    DATA_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
    THIN_BORDER = Border(
        left=Side(style="thin", color="E0E4E8"),
        right=Side(style="thin", color="E0E4E8"),
        top=Side(style="thin", color="E0E4E8"),
        bottom=Side(style="thin", color="E0E4E8")
    )

    def __init__(self):
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl 未安装，请执行: pip install openpyxl")

    def export_materials(self, data: List[Dict], filepath: Optional[str] = None) -> str:
        """
        导出物料台账
        参数:
            data: 物料数据列表
            filepath: 输出文件路径（可选，默认生成）
        返回:
            生成的文件路径
        """
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"物料台账_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "物料台账"

        # 表头定义 — 匹配实际 materials 表字段
        headers = [
            ("物料编码", "material_code"),
            ("物料名称", "material_name"),
            ("存放位置", "location"),
            ("货架号", "shelf_no"),
            ("库存数量", "stock_qty"),
            ("预留量", "reserved_qty"),
            ("单位", "unit"),
            ("最后更新", "last_update"),
        ]

        self._write_headers(ws, headers)
        self._write_data(ws, data, headers, start_row=2)
        self._auto_adjust_columns(ws, headers)

        wb.save(filepath)
        return filepath

    def export_inventory_records(self, data: List[Dict], filepath: Optional[str] = None) -> str:
        """导出领用/归还记录"""
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"领用记录_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "领用记录"

        headers = [
            ("记录单号", "record_no"),
            ("物料编码", "material_code"),
            ("物料名称", "material_name"),
            ("数量", "qty"),
            ("卡号", "card_no"),
            ("使用部门", "dept"),
            ("使用人", "user_name"),
            ("操作类型", "action_type"),
            ("出库时间", "out_time"),
            ("经办人", "operator"),
            ("是否归还", "is_returned"),
            ("归还时间", "in_time"),
            ("确认人", "confirm_person"),
            ("归还人", "return_person"),
            ("归还数量", "return_qty"),
            ("好板数", "good_qty"),
            ("坏板数", "damage_qty"),
            ("补单状态", "damage_status"),
            ("混板数量", "mixed_qty"),
            ("混板备注", "mixed_remark"),
        ]

        self._write_headers(ws, headers)
        self._write_data(ws, data, headers, start_row=2)
        self._auto_adjust_columns(ws, headers)

        wb.save(filepath)
        return filepath

    def export_check_records(self, data: List[Dict], filepath: Optional[str] = None) -> str:
        """导出盘点记录（基于 materials 表数据）"""
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"盘点记录_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "盘点记录"

        headers = [
            ("物料编码", "material_code"),
            ("物料名称", "material_name"),
            ("系统库存", "stock_qty"),
            ("存放位置", "location"),
            ("货架号", "shelf_no"),
            ("单位", "unit"),
            ("最后更新", "last_update"),
        ]

        self._write_headers(ws, headers)
        self._write_data(ws, data, headers, start_row=2)
        self._auto_adjust_columns(ws, headers)

        wb.save(filepath)
        return filepath

    def export_employees(self, data, filepath=None):
        """导出员工档案"""
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"员工档案_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "员工档案"

        headers = [
            ("工号", "employee_no"),
            ("姓名", "name"),
            ("部门", "dept"),
            ("联系电话", "phone"),
            ("指纹ID", "fingerprint_id"),
            ("NFC卡号", "card_no"),
            ("状态", "is_active"),
            ("创建时间", "created_at"),
            ("更新时间", "updated_at"),
        ]

        self._write_headers(ws, headers)
        self._write_data(ws, data, headers, start_row=2)
        self._auto_adjust_columns(ws, headers)

        wb.save(filepath)
        return filepath

    def export_alerts(self, data: List[Dict], filepath: Optional[str] = None) -> str:
        """导出低库存预警记录（基于 materials 表数据）"""
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"预警记录_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "预警记录"

        headers = [
            ("物料编码", "material_code"),
            ("物料名称", "material_name"),
            ("库存数量", "stock_qty"),
            ("存放位置", "location"),
            ("货架号", "shelf_no"),
            ("单位", "unit"),
            ("最后更新", "last_update"),
        ]

        self._write_headers(ws, headers)
        self._write_data(ws, data, headers, start_row=2)
        self._auto_adjust_columns(ws, headers)

        wb.save(filepath)
        return filepath

    def export_multi_sheet(self, sheets_data: Dict[str, List[Dict]], filepath: Optional[str] = None) -> str:
        """
        导出多 Sheet 综合报表
        参数:
            sheets_data: {sheet_name: data_list}
        """
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"综合报表_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        # 删除默认 sheet
        wb.remove(wb.active)

        for sheet_name, data in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)
            if data:
                headers = [(k, k) for k in data[0].keys()]
                self._write_headers(ws, headers)
                self._write_data(ws, data, headers, start_row=2)
                self._auto_adjust_columns(ws, headers)

        wb.save(filepath)
        return filepath

    def _write_headers(self, ws, headers: List[tuple]):
        """写入表头"""
        for col_idx, (header_text, _) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header_text)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.THIN_BORDER
        ws.row_dimensions[1].height = 30

    def _write_data(self, ws, data: List[Dict], headers: List[tuple], start_row: int = 2):
        """写入数据行"""
        for row_idx, record in enumerate(data, start_row):
            for col_idx, (_, key) in enumerate(headers, 1):
                value = record.get(key, "")
                # 处理布尔值
                if isinstance(value, bool):
                    value = "是" if value else "否"
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = self.DATA_FONT
                cell.alignment = self.DATA_ALIGNMENT
                cell.border = self.THIN_BORDER
                # 数值居中
                if isinstance(value, (int, float)):
                    cell.alignment = self.CENTER_ALIGNMENT
            ws.row_dimensions[row_idx].height = 24

    def _auto_adjust_columns(self, ws, headers: List[tuple]):
        """自动调整列宽"""
        for col_idx, (header_text, _) in enumerate(headers, 1):
            max_length = len(str(header_text))
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
                for cell in row:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    def export_inventory(self, data: List[Dict], filepath: Optional[str] = None) -> str:
        """导出库存明细"""
        if filepath is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filepath = f"库存明细_{timestamp}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "库存明细"

        headers = [
            ("地点", "location"),
            ("货架号", "shelf_no"),
            ("物料编码", "material_code"),
            ("物料名称", "material_name"),
            ("库存数量", "stock_qty"),
            ("单位", "unit"),
            ("预留量", "reserved_qty"),
            ("实物图", "real_image"),
            ("最后更新", "last_update"),
        ]

        self._write_headers(ws, headers)
        self._write_data(ws, data, headers, start_row=2)
        self._auto_adjust_columns(ws, headers)

        wb.save(filepath)
        return filepath

    @staticmethod
    def is_available() -> bool:
        """检查 openpyxl 是否可用"""
        return EXCEL_AVAILABLE
